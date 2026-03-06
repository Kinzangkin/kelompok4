import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from PIL import Image, ImageTk
import cv2
import numpy as np
import os
import math
from datetime import datetime
from components.selectors import CameraSelector

class ShapeAnalysisPanel:
    def __init__(self, parent_app):
        self.app = parent_app
        self.root = parent_app.root
        self.camera = None
        self.camera_running = False
        self.captured_image = None
        self.current_frame = None
        self.result_image = None
        self.current_stats = {}

    def show_selection(self):
        """Show camera source selection dialog"""
        selector = CameraSelector(
            self.root,
            "Pilih Sumber Kamera (Analisis Bentuk)",
            self.open_panel
        )
        selector.show()

    def open_panel(self, source):
        """Open the shape analysis panel"""
        self.shape_window = tk.Toplevel(self.root)
        self.shape_window.title("Analisis Bentuk")
        self.shape_window.geometry("1366x768")
        self.shape_window.configure(bg="#1a202c")
        self.shape_window.state('zoomed')
        self.shape_window.resizable(True, True)

        # =============== MAIN HORIZONTAL LAYOUT ===============
        main_hbox = tk.Frame(self.shape_window, bg="#1a202c")
        main_hbox.pack(fill=tk.BOTH, expand=True, padx=0, pady=0)

        # =============== LEFT SIDEBAR ===============
        self._build_sidebar(main_hbox)

        # =============== CENTER CONTENT AREA ===============
        content_area = tk.Frame(main_hbox, bg="#1a202c")
        content_area.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=15, pady=15)

        # --- IMAGE VIEWERS: Camera Input | Result ---
        images_frame = tk.Frame(content_area, bg="#1a202c")
        images_frame.pack(fill=tk.BOTH, expand=True)

        # Camera panel
        camera_frame = tk.Frame(images_frame, bg="#2d3748", relief=tk.FLAT, bd=0)
        camera_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))
        tk.Label(camera_frame, text="❂ Camera Input", font=("Helvetica", 11, "bold"),
                 bg="#2d3748", fg="white", anchor="w").pack(fill=tk.X, padx=15, pady=(15, 5))
        self.camera_canvas = tk.Canvas(camera_frame, bg="#1a202c", width=420, height=360, highlightthickness=0)
        self.camera_canvas.pack(padx=15, pady=(0, 15), fill=tk.BOTH, expand=True)

        # Timer Label (hidden by default)
        self.timer_label = tk.Label(self.camera_canvas, text="", font=("Helvetica", 48, "bold"),
                                   bg="#1a202c", fg="#ef4444")

        # Result panel
        result_frame = tk.Frame(images_frame, bg="#2d3748", relief=tk.FLAT, bd=0)
        result_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(10, 0))
        tk.Label(result_frame, text="🖻 Processed Result", font=("Helvetica", 11, "bold"),
                 bg="#2d3748", fg="white", anchor="w").pack(fill=tk.X, padx=15, pady=(15, 5))
        self.result_canvas = tk.Canvas(result_frame, bg="#1a202c", width=420, height=360, highlightthickness=0)
        self.result_canvas.pack(padx=15, pady=(0, 15), fill=tk.BOTH, expand=True)

        # --- THRESHOLD SLIDER ---
        slider_container = tk.Frame(content_area, bg="#2d3748", relief=tk.FLAT, bd=0)
        slider_container.pack(fill=tk.X, pady=(20, 0))
        
        slider_inner = tk.Frame(slider_container, bg="#2d3748")
        slider_inner.pack(fill=tk.X, padx=20, pady=15)

        tk.Label(slider_inner, text="Threshold (Otsu by default):", font=("Helvetica", 10, "bold"),
                 bg="#2d3748", fg="#d69e2e").pack(side=tk.LEFT)

        self.threshold_label = tk.Label(slider_inner, text="Auto", font=("Helvetica", 10, "bold"),
                                        bg="#2d3748", fg="lightgray", width=5)
        self.threshold_label.pack(side=tk.RIGHT)

        style = ttk.Style()
        style.configure("Modern.Horizontal.TScale", background="#2d3748", troughcolor="#4a5568")

        self.threshold_slider = ttk.Scale(
            slider_inner, from_=0, to=255, orient=tk.HORIZONTAL,
            command=self._on_threshold_change, style="Modern.Horizontal.TScale"
        )
        self.threshold_slider.set(127) # Default but will be overridden by otsu calculation later
        self.threshold_slider.pack(fill=tk.X, padx=15, expand=True)

        # --- BOTTOM ACTION BUTTONS ---
        btn_container = tk.Frame(content_area, bg="#1a202c")
        btn_container.pack(fill=tk.X, pady=(20, 0))
        
        btn_inner = tk.Frame(btn_container, bg="#1a202c")
        btn_inner.pack(anchor="center")

        btn_style = {"font": ("Helvetica", 11, "bold"), "width": 14, "height": 1, "cursor": "hand2", "relief": tk.FLAT, "pady": 8}

        self.save_btn = tk.Button(btn_inner, text="💾 Simpan",
                                  command=self.save_image,
                                  bg="#10b981", fg="white", activebackground="#059669", activeforeground="white",
                                  disabledforeground="white", state=tk.DISABLED, **btn_style)
        self.save_btn.pack(side=tk.LEFT, padx=15)

        self.export_btn = tk.Button(btn_inner, text="📄 Export Excel",
                                  command=self.export_excel,
                                  bg="#0ea5e9", fg="white", activebackground="#0284c7", activeforeground="white",
                                  disabledforeground="white", state=tk.DISABLED, **btn_style)
        self.export_btn.pack(side=tk.LEFT, padx=15)

        self.close_btn = tk.Button(btn_inner, text="✕ Tutup Aplikasi",
                                   command=self.close_panel,
                                   bg="#ef4444", fg="white", activebackground="#dc2626", activeforeground="white", **btn_style)
        self.close_btn.pack(side=tk.LEFT, padx=15)

        # =============== RIGHT SIDEBAR STATISTICS ===============
        self._build_right_sidebar(main_hbox)

        # Start camera
        self.shape_window.protocol("WM_DELETE_WINDOW", self.close_panel)
        self.start_camera(source)

    # =============== LAYOUT BUILDERS ===============

    def _build_sidebar(self, parent):
        """Build the left sidebar for actions"""
        sidebar = tk.Frame(parent, bg="#2d3748", width=220, relief=tk.FLAT)
        sidebar.pack(side=tk.LEFT, fill=tk.Y)
        sidebar.pack_propagate(False)

        # Header
        header = tk.Frame(sidebar, bg="#2d3748", pady=15)
        header.pack(fill=tk.X, padx=15)
        tk.Label(header, text="❖ Sumber Data", font=("Helvetica", 13, "bold"),
                 bg="#2d3748", fg="white", anchor="w").pack(fill=tk.X)
        
        tk.Frame(sidebar, bg="#4a5568", height=1).pack(fill=tk.X, padx=15, pady=(0, 10))

        # Base actions
        btn_style = {"font": ("Helvetica", 11, "bold"), "relief": tk.FLAT, "cursor": "hand2", "anchor": tk.W, "padx": 15, "pady": 10}
        
        self.btn_drive = tk.Button(sidebar, text="📁 Drive", command=self._sidebar_open,
                  bg="#3b82f6", fg="white", activebackground="#2563eb", activeforeground="white", **btn_style)
        self.btn_drive.pack(fill=tk.X, padx=15, pady=(5, 5))

        self.btn_camera = tk.Button(sidebar, text="📷 Camera", command=self._sidebar_camera,
                  bg="#6366f1", fg="white", activebackground="#4f46e5", activeforeground="white", **btn_style)
        self.btn_camera.pack(fill=tk.X, padx=15, pady=(5, 5))
        
        # Separator before capture
        tk.Frame(sidebar, bg="#2d3748", height=20).pack(fill=tk.X)

        self.capture_btn = tk.Button(sidebar, text="📸 Capture", command=self.capture_image,
                  bg="#22c55e", fg="white", activebackground="#16a34a", activeforeground="white", **btn_style)
        self.capture_btn.pack(fill=tk.X, padx=15, pady=(5, 5))

        self.timer_btn = tk.Button(sidebar, text="⏱️ Timer (3s)", command=self.capture_with_timer,
                  bg="#f59e0b", fg="white", activebackground="#d97706", activeforeground="white", **btn_style)
        self.timer_btn.pack(fill=tk.X, padx=15, pady=(5, 5))


    def _build_right_sidebar(self, parent):
        """Build the right sidebar for shape statistics"""
        sidebar = tk.Frame(parent, bg="#2d3748", width=260, relief=tk.FLAT)
        sidebar.pack(side=tk.RIGHT, fill=tk.Y)
        sidebar.pack_propagate(False)
        
        # Header
        header = tk.Frame(sidebar, bg="#2d3748", pady=15)
        header.pack(fill=tk.X, padx=15)
        tk.Label(header, text="📊 Parameter Bentuk", font=("Helvetica", 12, "bold"),
                 bg="#2d3748", fg="white", anchor="w").pack(fill=tk.X)

        tk.Frame(sidebar, bg="#4a5568", height=1).pack(fill=tk.X, padx=15)
        
        stats_frame = tk.Frame(sidebar, bg="#2d3748")
        stats_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)
        
        self.stat_labels = {}
        # Metrics: Area, Width, Length, Perimeter, Dispersion, Roundness, Slenderness, Convex Hull Area, Hull Keliling, Solidity
        stat_items = [
            ("Luas (Area)", "luas", "px²"),
            ("Lebar (Width)", "lebar", "px"),
            ("Panjang (Length)", "panjang", "px"),
            ("Perimeter", "perimeter", "px"),
            ("Dispersi", "dispersi", ""),
            ("Kebulatan", "kebulatan", ""),
            ("Kerampingan", "kerampingan", ""),
            ("Luas Convex Hull", "hull_area", "px²"),
            ("Keliling Hull", "hull_perimeter", "px"),
            ("Soliditas", "soliditas", ""),
        ]
        
        for idx, (label_text, key, unit) in enumerate(stat_items):
            tk.Label(stats_frame, text=label_text, font=("Helvetica", 10, "bold"),
                     bg="#2d3748", fg="#d69e2e", anchor=tk.W).pack(anchor=tk.W, pady=(15 if idx > 0 else 5, 0))
            
            val_frame = tk.Frame(stats_frame, bg="#2d3748")
            val_frame.pack(anchor=tk.W, fill=tk.X)
            
            val_label = tk.Label(val_frame, text="-", font=("Consolas", 11),
                                 bg="#2d3748", fg="lightgray", anchor=tk.W)
            val_label.pack(side=tk.LEFT, pady=(2, 0))
            
            if unit:
                tk.Label(val_frame, text=f" {unit}", font=("Helvetica", 9),
                         bg="#2d3748", fg="gray", anchor=tk.W).pack(side=tk.LEFT, pady=(4, 0))
                
            self.stat_labels[key] = val_label

    # =============== ACTIONS ===============

    def _sidebar_camera(self):
        """Switch back to live camera view"""
        if not self.camera_running and self.camera:
            self.camera_running = True
            self.captured_image = None
            self.capture_btn.config(text="📸 Capture", bg="#22c55e")
            self.export_btn.config(state=tk.DISABLED)
            self._clear_stats()
            self._update_camera_loop()
            self.app.update_status("Kamera live view - Analisis Bentuk")

    def _sidebar_open(self):
        """Open image from file"""
        filepath = filedialog.askopenfilename(
            initialdir=self.app.gallery_folder,
            title="Pilih Gambar",
            filetypes=[("Image files", "*.png *.jpg *.jpeg *.bmp"), ("All files", "*.*")]
        )
        if filepath:
            # Stop camera
            self.camera_running = False
            img_bgr = cv2.imread(filepath)
            if img_bgr is not None:
                self.captured_image = img_bgr.copy()
                self._show_on_canvas(self.camera_canvas, self.captured_image)
                self.capture_btn.config(text="🔄 Ambil Ulang", bg="#f59e0b", activebackground="#d97706")
                self.app.update_status(f"Gambar dimuat: {os.path.basename(filepath)}")
                self._analyze_shape() # Auto analyze

    def capture_image(self):
        """Capture or retake image from camera immediately"""
        if not self.camera_running:
            # RETAKE
            self.camera_running = True
            self.captured_image = None
            self.capture_btn.config(text="📸 Capture", bg="#22c55e", activebackground="#16a34a")
            self.timer_btn.config(state=tk.NORMAL)
            self.export_btn.config(state=tk.DISABLED)
            self.save_btn.config(state=tk.DISABLED)
            self._clear_stats()
            self._update_camera_loop()
            self.app.update_status("Kamera live view")
        else:
            # CAPTURE
            if self.camera and self.current_frame is not None:
                self.captured_image = self.current_frame.copy()
                self.camera_running = False
                self._show_on_canvas(self.camera_canvas, self.captured_image)
                self.capture_btn.config(text="🔄 Ambil Ulang", bg="#f59e0b", activebackground="#d97706")
                self.timer_btn.config(state=tk.DISABLED)
                self.app.update_status("Gambar di-capture, menghitung parameter bentuk...")
                self._analyze_shape()
            else:
                messagebox.showwarning("Peringatan", "Tidak ada frame kamera!")

    def capture_with_timer(self):
        """Start a 3-second countdown before capturing"""
        if not self.camera_running:
            return  # Can only use timer during live view

        # Disable buttons during countdown
        self.btn_drive.config(state=tk.DISABLED)
        self.btn_camera.config(state=tk.DISABLED)
        self.capture_btn.config(state=tk.DISABLED)
        self.timer_btn.config(state=tk.DISABLED)
        self.export_btn.config(state=tk.DISABLED)
        self.save_btn.config(state=tk.DISABLED)
        
        self.app.update_status("Timer dimulai...")
        self._countdown(3)

    def _countdown(self, seconds_left):
        """Recursive function to handle the timer UI updates"""
        if seconds_left > 0:
            # Update label text and ensure it is placed on canvas
            self.timer_label.config(text=str(seconds_left))
            self.timer_label.place(relx=0.5, rely=0.5, anchor=tk.CENTER)
            self.shape_window.after(1000, self._countdown, seconds_left - 1)
        else:
            # Remove timer label and capture
            self.timer_label.place_forget()
            
            # Re-enable buttons
            self.btn_drive.config(state=tk.NORMAL)
            self.btn_camera.config(state=tk.NORMAL)
            self.capture_btn.config(state=tk.NORMAL)
            self.timer_btn.config(state=tk.NORMAL)
            
            # Trigger actual capture
            self.capture_image()

    def _on_threshold_change(self, val):
        """Handle manual threshold change"""
        if self.captured_image is not None:
            self.threshold_label.config(text=str(int(float(val))))
            self._analyze_shape(manual_threshold=int(float(val)))

    # =============== CAMERA SHAPE ANALYSIS ===============

    def start_camera(self, source):
        try:
            self.camera = cv2.VideoCapture(source)
            if not self.camera.isOpened():
                messagebox.showerror("Error", "Gagal membuka kamera!")
                self.shape_window.destroy()
                return
            self.camera_running = True
            self.app.update_status("Kamera aktif - Mode Analisis Bentuk")
            self._update_camera_loop()
        except Exception as e:
            messagebox.showerror("Error", f"Error: {str(e)}")
            self.shape_window.destroy()

    def _update_camera_loop(self):
        if self.camera_running and self.camera:
            ret, frame = self.camera.read()
            if ret:
                self.current_frame = frame.copy()
                self._show_on_canvas(self.camera_canvas, frame)
            try:
                self.shape_window.after(30, self._update_camera_loop)
            except tk.TclError:
                pass

    def _analyze_shape(self, manual_threshold=None):
        """Process image, find contours, and calculate shape metrics"""
        if self.captured_image is None:
            return

        # 1. Grayscale & Blur
        gray = cv2.cvtColor(self.captured_image, cv2.COLOR_BGR2GRAY)
        blurred = cv2.GaussianBlur(gray, (5, 5), 0)

        # 2. Thresholding
        if manual_threshold is None:
            ret, thresh = cv2.threshold(blurred, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
            # Update slider to reflect Otsu value
            self.threshold_slider.set(ret)
            self.threshold_label.config(text=f"Auto ({int(ret)})")
        else:
            _, thresh = cv2.threshold(blurred, manual_threshold, 255, cv2.THRESH_BINARY_INV)

        # Morphological operations to remove noise
        kernel = np.ones((3,3), np.uint8)
        opening = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, kernel, iterations=2)
        
        # 3. Find Contours
        contours, hierarchy = cv2.findContours(opening, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        
        display_img = self.captured_image.copy()

        if contours:
            # Find the largest contour assuming it's the main object
            main_contour = max(contours, key=cv2.contourArea)
            
            # Calculate Convex Hull
            hull = cv2.convexHull(main_contour)
            
            # Draw contour (green) and convex hull (red)
            cv2.drawContours(display_img, [main_contour], -1, (0, 255, 0), 2)
            cv2.drawContours(display_img, [hull], -1, (0, 0, 255), 2)
            
            # Calculate metrics
            area = cv2.contourArea(main_contour)
            perimeter = cv2.arcLength(main_contour, True)
            hull_area = cv2.contourArea(hull)
            hull_perimeter = cv2.arcLength(hull, True)
            
            # Bounding Box for Width and Length
            x, y, w, h = cv2.boundingRect(main_contour)
            cv2.rectangle(display_img, (x, y), (x+w, y+h), (255, 0, 0), 2) # Draw bounding box in blue
            
            lebar = w
            panjang = h
            
            # Derived Metrics
            dispersi = (perimeter ** 2) / area if area > 0 else 0
            kebulatan = (4 * math.pi * area) / (perimeter ** 2) if perimeter > 0 else 0
            kerampingan = panjang / lebar if lebar > 0 else 0
            soliditas = area / hull_area if hull_area > 0 else 0
            
            # Update dictionary
            self.current_stats = {
                "Luas": area,
                "Lebar": lebar,
                "Panjang": panjang,
                "Perimeter": perimeter,
                "Dispersi": dispersi,
                "Kebulatan": kebulatan,
                "Kerampingan": kerampingan,
                "Hull_Area": hull_area,
                "Hull_Perimeter": hull_perimeter,
                "Soliditas": soliditas,
                "Unit_Px": True # indicator for excel export
            }
            
            # Update UI Labels
            self.stat_labels['luas'].config(text=f"{area:,.1f}")
            self.stat_labels['lebar'].config(text=f"{lebar:,.1f}")
            self.stat_labels['panjang'].config(text=f"{panjang:,.1f}")
            self.stat_labels['perimeter'].config(text=f"{perimeter:,.2f}")
            self.stat_labels['dispersi'].config(text=f"{dispersi:,.4f}")
            self.stat_labels['kebulatan'].config(text=f"{kebulatan:,.4f}")
            self.stat_labels['kerampingan'].config(text=f"{kerampingan:,.4f}")
            self.stat_labels['hull_area'].config(text=f"{hull_area:,.1f}")
            self.stat_labels['hull_perimeter'].config(text=f"{hull_perimeter:,.2f}")
            self.stat_labels['soliditas'].config(text=f"{soliditas:,.4f}")
            
            self.export_btn.config(state=tk.NORMAL)
            self.save_btn.config(state=tk.NORMAL)
            self.app.update_status("Analisis selesai. Parameter bentuk berhasil dihitung.")

        else:
            self._clear_stats()
            self.app.update_status("Objek tidak terdeteksi. Silakan atur threshold.")

        self.result_image = display_img
        self._show_on_canvas(self.result_canvas, self.result_image)


    def _clear_stats(self):
        for key in self.stat_labels:
            self.stat_labels[key].config(text="-")
        self.result_canvas.delete("all")
        self.current_stats = {}

    def save_image(self):
        """Save the processed result image to gallery"""
        if self.result_image is None:
            messagebox.showwarning("Peringatan", "Tidak ada hasil proses untuk disimpan!")
            return
            
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"shape_analysis_{timestamp}.png"
        filepath = os.path.join(self.app.gallery_folder, filename)
        
        try:
            cv2.imwrite(filepath, self.result_image)
            self.app.update_status(f"Gambar disimpan: {filename}")
            messagebox.showinfo("Sukses", f"Gambar berhasil disimpan ke:\n{filepath}")
        except Exception as e:
            messagebox.showerror("Error", f"Gagal menyimpan gambar: {str(e)}")

    def export_excel(self):
        """Export current shape statistics to Excel file"""
        if not self.current_stats:
            messagebox.showwarning("Peringatan", "Tidak ada data untuk di-export!")
            return
            
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            messagebox.showerror("Error", "Module 'openpyxl' belum terinstall!\nJalankan: pip install openpyxl")
            return

        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            initialfile=f"Shape_Analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        if not filepath:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Shape Image Stats"

        # Headers
        headers = ["Parameter", "Nilai", "Satuan"]
        ws.append(headers)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4A5568", end_color="4A5568", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        s = self.current_stats
        data_rows = [
            ["Luas (Area)", s.get("Luas", 0), "px²"],
            ["Lebar (Width)", s.get("Lebar", 0), "px"],
            ["Panjang (Length)", s.get("Panjang", 0), "px"],
            ["Perimeter / Keliling", s.get("Perimeter", 0), "px"],
            ["Dispersi", s.get("Dispersi", 0), ""],
            ["Kebulatan", s.get("Kebulatan", 0), ""],
            ["Kerampingan", s.get("Kerampingan", 0), ""],
            ["Luas Convex Hull", s.get("Hull_Area", 0), "px²"],
            ["Keliling Hull", s.get("Hull_Perimeter", 0), "px"],
            ["Soliditas", s.get("Soliditas", 0), ""]
        ]

        for row in data_rows:
            ws.append(row)

        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max(max_length + 2, 15)

        wb.save(filepath)
        self.app.update_status(f"Export berhasil: {os.path.basename(filepath)}")
        messagebox.showinfo("Sukses", f"Data berhasil di-export ke:\n{filepath}")

    # =============== UTILITY ===============

    def _show_on_canvas(self, canvas, img, is_gray=False):
        """Display an image on a tk.Canvas, fitting to canvas size"""
        try:
            canvas.update_idletasks()
            cw = canvas.winfo_width() or 420
            ch = canvas.winfo_height() or 360
        except tk.TclError:
            cw, ch = 420, 360
            
        # Safety check to prevent division by zero or negative sizes
        if cw <= 0: cw = 420
        if ch <= 0: ch = 360

        if is_gray:
            img_rgb = img
        else:
            img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)

        # Maintain aspect ratio
        h, w = img_rgb.shape[:2]
        
        # Prevent division by zero
        if h == 0 or ch == 0:
            return

        display_ratio = cw / ch
        img_ratio = w / h
        
        if img_ratio > display_ratio:
            new_w = cw
            new_h = int(cw / img_ratio)
        else:
            new_h = ch
            new_w = int(ch * img_ratio)
            
        # Ensure dimensions are strictly positive before resizing
        if new_w <= 0: new_w = 1
        if new_h <= 0: new_h = 1
            
        resized_img = cv2.resize(img_rgb, (new_w, new_h))
        pil_img = Image.fromarray(resized_img)
        tk_img = ImageTk.PhotoImage(pil_img)

        canvas.delete("all")
        canvas.create_image(cw // 2, ch // 2, image=tk_img)
        canvas._tk_img_ref = tk_img  # prevent GC

    def close_panel(self):
        """Close the shape analysis panel"""
        self.camera_running = False
        if self.camera:
            self.camera.release()
            self.camera = None
        self.shape_window.destroy()
        self.app.show_gallery_page()
