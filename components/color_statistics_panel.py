import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from PIL import Image, ImageTk
import cv2
import numpy as np
import os
import sqlite3
from datetime import datetime
from components.selectors import CameraSelector
from components.utils import get_color_name

# Try importing matplotlib for histogram chart
import matplotlib
matplotlib.use('TkAgg')
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure

class ColorStatisticsPanel:
    def __init__(self, parent_app):
        self.app = parent_app
        self.root = parent_app.root
        self.camera = None
        self.camera_running = False
        self.captured_frame = None
        self.db_path = os.path.join(self.app.app_dir, "color_stats.db")
        self._init_database()

    def _init_database(self):
        """Initialize SQLite database for color statistics"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS color_stats (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                mean_r REAL, mean_g REAL, mean_b REAL,
                unique_colors INTEGER,
                dominant_hex TEXT,
                dominant_name TEXT,
                timestamp TEXT
            )
        ''')
        conn.commit()
        conn.close()

    def show_selection(self):
        """Show camera source selection dialog"""
        selector = CameraSelector(
            self.root,
            "Pilih Sumber Kamera (Statistik Warna)",
            self.open_panel
        )
        selector.show()

    def open_panel(self, source):
        """Open the color statistics panel"""
        self.color_window = tk.Toplevel(self.root)
        self.color_window.title("Statistik Warna Kamera")
        self.color_window.geometry("1200x800")
        self.color_window.configure(bg="#1a202c")
        self.color_window.state('zoomed')
        
        # Main container with 2 columns
        main_container = tk.Frame(self.color_window, bg="#1a202c")
        main_container.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # LEFT COLUMN (Camera & Info)
        left_col = tk.Frame(main_container, bg="#1a202c")
        left_col.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))
        
        # Camera Feed
        camera_frame = tk.Frame(left_col, bg="#2d3748", relief=tk.FLAT, bd=0)
        camera_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        tk.Label(camera_frame, text="📸 LIVE CAMERA FEED", font=("Helvetica", 12, "bold"),
                 bg="#2d3748", fg="white").pack(pady=10)
        
        self.camera_canvas = tk.Canvas(camera_frame, bg="#1a252f", highlightthickness=0)
        self.camera_canvas.pack(fill=tk.BOTH, expand=True, padx=15, pady=(0, 15))
        
        # Info Panel (Stats)
        stats_frame = tk.Frame(left_col, bg="#2d3748", pady=15, padx=20)
        stats_frame.pack(fill=tk.X)
        
        # Grid for stats
        stats_grid = tk.Frame(stats_frame, bg="#2d3748")
        stats_grid.pack(fill=tk.X)
        
        # Mean RGB
        tk.Label(stats_grid, text="Rata-rata RGB:", font=("Helvetica", 11, "bold"),
                 bg="#2d3748", fg="#cbd5e0").grid(row=0, column=0, sticky="w", pady=5)
        self.mean_label = tk.Label(stats_grid, text="R: 0, G: 0, B: 0", font=("Consolas", 12, "bold"),
                                   bg="#2d3748", fg="#f6e05e")
        self.mean_label.grid(row=0, column=1, sticky="w", padx=20)
        
        # Unique Colors
        tk.Label(stats_grid, text="Jumlah Warna Unik:", font=("Helvetica", 11, "bold"),
                 bg="#2d3748", fg="#cbd5e0").grid(row=1, column=0, sticky="w", pady=5)
        self.unique_label = tk.Label(stats_grid, text="0", font=("Consolas", 12, "bold"),
                                     bg="#2d3748", fg="#63b3ed")
        self.unique_label.grid(row=1, column=1, sticky="w", padx=20)
        
        # Dominant Color
        tk.Label(stats_grid, text="Warna Dominan:", font=("Helvetica", 11, "bold"),
                 bg="#2d3748", fg="#cbd5e0").grid(row=2, column=0, sticky="w", pady=5)
        
        dom_color_container = tk.Frame(stats_grid, bg="#2d3748")
        dom_color_container.grid(row=2, column=1, sticky="w", padx=20)
        
        self.dominant_box = tk.Canvas(dom_color_container, width=50, height=30, bg="black", highlightthickness=1, highlightbackground="white")
        self.dominant_box.pack(side=tk.LEFT)
        
        self.dominant_label = tk.Label(dom_color_container, text="#000000 (Black)", font=("Helvetica", 11, "bold"),
                                       bg="#2d3748", fg="white")
        self.dominant_label.pack(side=tk.LEFT, padx=10)
        
        # RIGHT COLUMN (Histogram & Buttons)
        right_col = tk.Frame(main_container, bg="#1a202c", width=450)
        right_col.pack(side=tk.RIGHT, fill=tk.BOTH, padx=(10, 0))
        right_col.pack_propagate(False)
        
        # Histogram
        hist_frame = tk.Frame(right_col, bg="#2d3748", relief=tk.FLAT, bd=0)
        hist_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 20))
        
        tk.Label(hist_frame, text="📊 REAL-TIME HISTOGRAM RGB", font=("Helvetica", 12, "bold"),
                 bg="#2d3748", fg="white").pack(pady=10)
        
        self.fig = Figure(figsize=(5, 4), dpi=100)
        self.fig.patch.set_facecolor('#1a202c')
        self.ax = self.fig.add_subplot(111)
        self.ax.set_facecolor('#1a202c')
        self.ax.tick_params(colors='white', labelsize=8)
        self.ax.spines['bottom'].set_color('white')
        self.ax.spines['left'].set_color('white')
        self.ax.spines['top'].set_visible(False)
        self.ax.spines['right'].set_visible(False)
        self.ax.set_xlabel('Intensitas', color='white', fontsize=9)
        self.ax.set_ylabel('Frekuensi', color='white', fontsize=9)
        
        self.chart_canvas = FigureCanvasTkAgg(self.fig, master=hist_frame)
        self.chart_canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))
        
        # Buttons
        btn_container = tk.Frame(right_col, bg="#1a202c")
        btn_container.pack(fill=tk.X)
        
        btn_style = {"font": ("Helvetica", 11, "bold"), "height": 2, "cursor": "hand2", "relief": tk.FLAT, "pady": 10}
        
        self.save_btn = tk.Button(btn_container, text="💾 SIMPAN KE DATABASE",
                                 command=self.save_to_database,
                                 bg="#10b981", fg="white", activebackground="#059669", activeforeground="white", **btn_style)
        self.save_btn.pack(fill=tk.X, pady=5)
        
        self.export_btn = tk.Button(btn_container, text="📄 EXPORT KE EXCEL",
                                   command=self.export_to_excel,
                                   bg="#3b82f6", fg="white", activebackground="#2563eb", activeforeground="white", **btn_style)
        self.export_btn.pack(fill=tk.X, pady=5)
        
        self.close_btn = tk.Button(btn_container, text="✕ TUTUP",
                                  command=self.close_panel,
                                  bg="#ef4444", fg="white", activebackground="#dc2626", activeforeground="white", **btn_style)
        self.close_btn.pack(fill=tk.X, pady=5)
        
        # Start processing
        self.color_window.protocol("WM_DELETE_WINDOW", self.close_panel)
        self.start_camera(source)

    def start_camera(self, source):
        try:
            self.camera = cv2.VideoCapture(source)
            if not self.camera.isOpened():
                messagebox.showerror("Error", "Gagal membuka kamera!")
                self.color_window.destroy()
                return
            self.camera_running = True
            self.app.update_status("Kamera aktif - Statistik Warna Real-time")
            self.update_loop()
        except Exception as e:
            messagebox.showerror("Error", f"Error: {str(e)}")
            self.color_window.destroy()

    def update_loop(self):
        if self.camera_running and self.camera:
            ret, frame = self.camera.read()
            if ret:
                self.captured_frame = frame.copy()
                
                # Update Camera Display
                self._show_on_canvas(self.camera_canvas, frame)
                
                # Calculate Color Stats
                self._process_color_stats(frame)
                
            self.color_window.after(30, self.update_loop)

    def _process_color_stats(self, frame):
        """Calculate and update color statistics in real-time"""
        # 1. Mean RGB
        # OpenCV uses BGR
        mean_bgr = cv2.mean(frame)[:3]
        mean_r, mean_g, mean_b = int(mean_bgr[2]), int(mean_bgr[1]), int(mean_bgr[0])
        self.mean_label.config(text=f"R: {mean_r}, G: {mean_g}, B: {mean_b}")
        
        # 2. Unique Colors
        # Downsample for faster calculation
        small_frame = cv2.resize(frame, (100, 100))
        unique_colors = len(np.unique(small_frame.reshape(-1, 3), axis=0))
        self.unique_label.config(text=f"{unique_colors:,}")
        
        # 3. Dominant Color
        # Simple method: most frequent color in the downsampled image
        pixels = small_frame.reshape(-1, 3)
        # Use a dictionary or counts to find most frequent. 
        # For simplicity and speed, let's use a small k-means or just find max frequency of quantized colors
        # Let's use quantization to 8 colors per channel
        quantized = (pixels // 32) * 32
        unique_rows, counts = np.unique(quantized, axis=0, return_counts=True)
        dominant_bgr = unique_rows[np.argmax(counts)]
        dom_r, dom_g, dom_b = int(dominant_bgr[2]), int(dominant_bgr[1]), int(dominant_bgr[0])
        
        color_name, hex_color = get_color_name(dom_r, dom_g, dom_b)
        self.dominant_box.config(bg=hex_color)
        self.dominant_label.config(text=f"{hex_color} ({color_name})")
        
        # Save current stats for database/excel
        self.current_stats = {
            'mean_r': mean_r, 'mean_g': mean_g, 'mean_b': mean_b,
            'unique_colors': unique_colors,
            'dominant_hex': hex_color,
            'dominant_name': color_name
        }
        
        # 4. Histogram
        self._update_histogram(frame)

    def _update_histogram(self, frame):
        """Update the RGB Histogram canvas"""
        self.ax.clear()
        self.ax.set_facecolor('#1a202c')
        
        # Calculate histograms for each channel
        # frame is BGR
        color = ('b', 'g', 'r')
        for i, col in enumerate(color):
            hist = cv2.calcHist([frame], [i], None, [256], [0, 256])
            self.ax.plot(hist, color=col)
            self.ax.fill_between(range(256), hist.flatten(), color=col, alpha=0.2)
            
        self.ax.tick_params(colors='white', labelsize=8)
        self.ax.spines['bottom'].set_color('white')
        self.ax.spines['left'].set_color('white')
        self.ax.set_xlim([0, 256])
        
        self.fig.tight_layout()
        self.chart_canvas.draw_idle()

    def _show_on_canvas(self, canvas, img):
        """Display an image on a tk.Canvas, fitting to canvas size"""
        try:
            canvas.update_idletasks()
            cw = canvas.winfo_width()
            ch = canvas.winfo_height()
        except tk.TclError:
            cw, ch = 640, 480
        
        if cw <= 1 or ch <= 1: 
            cw, ch = 640, 480

        img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        h, w = img_rgb.shape[:2]

        display_ratio = cw / ch
        img_ratio = w / h
        
        if img_ratio > display_ratio:
            new_w = cw
            new_h = int(cw / img_ratio)
        else:
            new_h = ch
            new_w = int(ch * img_ratio)
            
        if new_w <= 0: new_w = 1
        if new_h <= 0: new_h = 1
            
        resized_img = cv2.resize(img_rgb, (new_w, new_h))
        pil_img = Image.fromarray(resized_img)
        tk_img = ImageTk.PhotoImage(pil_img)

        canvas.delete("all")
        canvas.create_image(cw // 2, ch // 2, image=tk_img)
        canvas._tk_img_ref = tk_img 

    def save_to_database(self):
        """Save captured metrics to database"""
        if not hasattr(self, 'current_stats'):
            return
            
        try:
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            s = self.current_stats
            
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO color_stats 
                (mean_r, mean_g, mean_b, unique_colors, dominant_hex, dominant_name, timestamp)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (s['mean_r'], s['mean_g'], s['mean_b'], s['unique_colors'], 
                  s['dominant_hex'], s['dominant_name'], timestamp))
            conn.commit()
            conn.close()
            
            self.app.update_status(f"Statistik warna disimpan ke database.")
            messagebox.showinfo("Sukses", "Data statistik warna berhasil disimpan ke database!")
        except Exception as e:
            messagebox.showerror("Error", f"Gagal menyimpan ke database: {e}")

    def export_to_excel(self):
        """Export database to Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            messagebox.showerror("Error", "Module 'openpyxl' belum terpasang!\nJalankan: pip install openpyxl")
            return

        # Fetch data from DB
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM color_stats ORDER BY id DESC')
        rows = cursor.fetchall()
        conn.close()

        if not rows:
            # If DB is empty, just export current frame if available
            if not hasattr(self, 'current_stats'):
                messagebox.showwarning("Peringatan", "Tidak ada data untuk diekspor!")
                return
            data_to_export = [(1, self.current_stats['mean_r'], self.current_stats['mean_g'], self.current_stats['mean_b'], 
                               self.current_stats['unique_colors'], self.current_stats['dominant_hex'], 
                               self.current_stats['dominant_name'], datetime.now().strftime("%Y-%m-%d %H:%M:%S"))]
        else:
            data_to_export = rows

        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            initialfile=f"Color_Stats_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        if not filepath:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Color Statistics"

        headers = ["ID", "Mean R", "Mean G", "Mean B", "Unique Colors", "Dominant Hex", "Dominant Name", "Timestamp"]
        ws.append(headers)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1A202C", end_color="1A202C", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row in data_to_export:
            ws.append(row)

        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max(max_length + 2, 12)

        wb.save(filepath)
        self.app.update_status(f"Export Excel berhasil: {os.path.basename(filepath)}")
        messagebox.showinfo("Sukses", f"Data berhasil di-export ke:\n{filepath}")

    def close_panel(self):
        """Close the panel and release resources"""
        self.camera_running = False
        if self.camera:
            self.camera.release()
            self.camera = None
        self.color_window.destroy()
        self.app.show_gallery_page()
