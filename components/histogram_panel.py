import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from PIL import Image, ImageTk
import cv2
import numpy as np
import os
import sqlite3
from datetime import datetime
from scipy import stats
from components.selectors import CameraSelector

# Try importing matplotlib for histogram chart
import matplotlib
matplotlib.use('TkAgg')
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure


class HistogramPanel:
    def __init__(self, parent_app):
        self.app = parent_app
        self.root = parent_app.root
        self.camera = None
        self.camera_running = False
        self.captured_image = None
        self.current_frame = None
        self.db_path = os.path.join(self.app.app_dir, "histogram_data.db")
        self._init_database()

    def _init_database(self):
        """Initialize SQLite database for histogram data"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS histogram_data (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nama_citra TEXT NOT NULL,
                skewness_r REAL, skewness_g REAL, skewness_b REAL,
                average_r REAL, average_g REAL, average_b REAL,
                std_r REAL, std_g REAL, std_b REAL,
                kurtosis_r REAL, kurtosis_g REAL, kurtosis_b REAL,
                timestamp TEXT
            )
        ''')
        conn.commit()
        conn.close()

    def show_selection(self):
        """Show camera source selection dialog"""
        selector = CameraSelector(
            self.root,
            "Pilih Sumber Kamera (Histogram)",
            self.open_panel
        )
        selector.show()

    def open_panel(self, source):
        """Open the histogram analysis panel"""
        self.hist_window = tk.Toplevel(self.root)
        self.hist_window.title("Analisis Histogram Citra")
        self.hist_window.geometry("1366x768")
        self.hist_window.configure(bg="#2c3e50")
        self.hist_window.state('zoomed')
        self.hist_window.resizable(True, True)

        # === TOP SECTION: Camera + Histogram + Stats ===
        top_frame = tk.Frame(self.hist_window, bg="#2c3e50")
        top_frame.pack(fill=tk.X, padx=15, pady=(10, 5))

        # --- Left: Camera / Citra ---
        camera_frame = tk.Frame(top_frame, bg="#34495e", relief=tk.SUNKEN, bd=2)
        camera_frame.pack(side=tk.LEFT, fill=tk.BOTH, padx=(0, 5))
        tk.Label(camera_frame, text="📷 Citra", font=("Helvetica", 12, "bold"),
                 bg="#34495e", fg="white").pack(pady=5)
        self.camera_canvas = tk.Canvas(camera_frame, bg="#1a252f", width=480, height=360)
        self.camera_canvas.pack(padx=8, pady=(0, 8))

        # --- Middle: Histogram Chart ---
        hist_chart_frame = tk.Frame(top_frame, bg="#34495e", relief=tk.SUNKEN, bd=2)
        hist_chart_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5)
        tk.Label(hist_chart_frame, text="📊 Histogram RGB", font=("Helvetica", 12, "bold"),
                 bg="#34495e", fg="white").pack(pady=5)

        # Matplotlib figure for histogram (Larger for fullscreen)
        self.fig = Figure(figsize=(8, 5), dpi=100)
        self.fig.patch.set_facecolor('#1a252f')
        self.ax = self.fig.add_subplot(111)
        self.ax.set_facecolor('#1a252f')
        self.ax.tick_params(colors='white', labelsize=7)
        self.ax.spines['bottom'].set_color('white')
        self.ax.spines['left'].set_color('white')
        self.ax.spines['top'].set_visible(False)
        self.ax.spines['right'].set_visible(False)
        self.ax.set_xlabel('Intensitas', color='white', fontsize=8)
        self.ax.set_ylabel('Frekuensi', color='white', fontsize=8)
        self.ax.text(0.5, 0.5, 'Klik "Histogram" untuk\nmenampilkan grafik',
                     transform=self.ax.transAxes, ha='center', va='center',
                     color='gray', fontsize=10)

        self.chart_canvas = FigureCanvasTkAgg(self.fig, master=hist_chart_frame)
        self.chart_canvas.draw()
        self.chart_canvas.get_tk_widget().pack(padx=8, pady=(0, 8))

        # Color channel checkboxes
        checkbox_frame = tk.Frame(hist_chart_frame, bg="#34495e")
        checkbox_frame.pack(pady=(0, 5))
        self.show_r = tk.BooleanVar(value=True)
        self.show_g = tk.BooleanVar(value=True)
        self.show_b = tk.BooleanVar(value=True)
        self.chk_r_btn = tk.Checkbutton(checkbox_frame, text="R", variable=self.show_r, bg="#34495e",
                       fg="#e74c3c", selectcolor="#2c3e50", font=("Helvetica", 9, "bold"),
                       activebackground="#34495e", command=self._refresh_histogram)
        self.chk_r_btn.pack(side=tk.LEFT, padx=5)
        
        self.chk_g_btn = tk.Checkbutton(checkbox_frame, text="G", variable=self.show_g, bg="#34495e",
                       fg="#2ecc71", selectcolor="#2c3e50", font=("Helvetica", 9, "bold"),
                       activebackground="#34495e", command=self._refresh_histogram)
        self.chk_g_btn.pack(side=tk.LEFT, padx=5)
        
        self.chk_b_btn = tk.Checkbutton(checkbox_frame, text="B", variable=self.show_b, bg="#34495e",
                       fg="#3498db", selectcolor="#2c3e50", font=("Helvetica", 9, "bold"),
                       activebackground="#34495e", command=self._refresh_histogram)
        self.chk_b_btn.pack(side=tk.LEFT, padx=5)

        # --- Right: Statistik Citra (Smaller) ---
        stats_frame = tk.Frame(top_frame, bg="#34495e", relief=tk.SUNKEN, bd=2, width=220)
        stats_frame.pack(side=tk.RIGHT, fill=tk.Y, padx=(5, 0))
        stats_frame.pack_propagate(False)
        tk.Label(stats_frame, text="📋 Statistik Citra", font=("Helvetica", 12, "bold"),
                 bg="#34495e", fg="white").pack(pady=5)

        info_inner = tk.Frame(stats_frame, bg="#1a252f", relief=tk.SUNKEN, bd=1)
        info_inner.pack(fill=tk.BOTH, expand=True, padx=8, pady=(0, 8))

        self.stat_labels = {}
        stat_items = [
            ("Info Citra", "info"),
            ("Skewness", "skewness"),
            ("Average", "average"),
            ("STD", "std"),
            ("Kurtosis", "kurtosis"),
        ]
        for label_text, key in stat_items:
            row = tk.Frame(info_inner, bg="#1a252f")
            row.pack(fill=tk.X, padx=10, pady=4)
            tk.Label(row, text=f"{label_text}:", font=("Helvetica", 9, "bold"),
                     bg="#1a252f", fg="#f1c40f", anchor=tk.W).pack(anchor=tk.W)
            val_label = tk.Label(row, text="-", font=("Consolas", 8),
                                 bg="#1a252f", fg="white", anchor=tk.W, wraplength=180, justify=tk.LEFT)
            val_label.pack(anchor=tk.W)
            self.stat_labels[key] = val_label

        # === MIDDLE SECTION: Database Table ===
        db_frame = tk.LabelFrame(self.hist_window, text=" 🗃️ Database ",
                                 font=("Helvetica", 11, "bold"), bg="#2c3e50",
                                 fg="white", relief=tk.GROOVE, bd=2)
        db_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=5)

        # Treeview table (Smaller height)
        columns = ("no", "nama_citra", "skewness", "average", "std", "kurtosis")
        self.tree = ttk.Treeview(db_frame, columns=columns, show="headings", height=4)

        # Style for treeview
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview",
                         background="#1a252f",
                         foreground="white",
                         fieldbackground="#1a252f",
                         font=("Helvetica", 9),
                         rowheight=25)
        style.configure("Treeview.Heading",
                         background="#34495e",
                         foreground="white",
                         font=("Helvetica", 10, "bold"))
        style.map("Treeview", background=[("selected", "#2980b9")])

        self.tree.heading("no", text="No")
        self.tree.heading("nama_citra", text="Nama Citra")
        self.tree.heading("skewness", text="Skewness")
        self.tree.heading("average", text="Average")
        self.tree.heading("std", text="STD")
        self.tree.heading("kurtosis", text="Kurtosis")

        self.tree.column("no", width=40, anchor=tk.CENTER)
        self.tree.column("nama_citra", width=200, anchor=tk.W)
        self.tree.column("skewness", width=200, anchor=tk.CENTER)
        self.tree.column("average", width=200, anchor=tk.CENTER)
        self.tree.column("std", width=200, anchor=tk.CENTER)
        self.tree.column("kurtosis", width=200, anchor=tk.CENTER)

        scrollbar = ttk.Scrollbar(db_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(5, 0), pady=5)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y, padx=(0, 5), pady=5)

        # Bind selection event
        self.tree.bind("<<TreeviewSelect>>", self._on_tree_select)

        # === BOTTOM SECTION: Buttons ===
        btn_frame = tk.Frame(self.hist_window, bg="#2c3e50")
        btn_frame.pack(fill=tk.X, padx=10, pady=(5, 12))

        btn_style = {"font": ("Helvetica", 15, "bold"), "width": 14, "height": 2, "cursor": "hand2"}

        self.capture_btn = tk.Button(btn_frame, text="📷 Capture",
                                     command=self.capture_image,
                                     bg="#2ecc71", fg="white", **btn_style)
        self.capture_btn.pack(side=tk.LEFT, padx=20, expand=True)

        self.histogram_btn = tk.Button(btn_frame, text="📊 Histogram",
                                       command=self.show_histogram,
                                       bg="#9b59b6", fg="white",
                                       state=tk.DISABLED, **btn_style)
        self.histogram_btn.pack(side=tk.LEFT, padx=20, expand=True)

        self.save_btn = tk.Button(btn_frame, text="💾 Simpan",
                                  command=self.save_to_database,
                                  bg="#3498db", fg="white",
                                  state=tk.DISABLED, **btn_style)
        self.save_btn.pack(side=tk.LEFT, padx=4, expand=True)

        self.export_btn = tk.Button(btn_frame, text="📄 Export Excel",
                                    command=self.export_to_excel,
                                    bg="#f39c12", fg="white", **btn_style)
        self.export_btn.pack(side=tk.LEFT, padx=4, expand=True)

        self.delete_btn = tk.Button(btn_frame, text="🗑️ Hapus",
                                    command=self.delete_selected,
                                    bg="#e74c3c", fg="white", **btn_style)
        self.delete_btn.pack(side=tk.LEFT, padx=4, expand=True)

        self.close_btn = tk.Button(btn_frame, text="❌ Tutup",
                                   command=self.close_panel,
                                   bg="#95a5a6", fg="white", **btn_style)
        self.close_btn.pack(side=tk.LEFT, padx=4, expand=True)

        # Start camera and load database
        self.start_camera(source)
        self._load_database()
        self.hist_window.protocol("WM_DELETE_WINDOW", self.close_panel)

    # ============== Camera Methods ==============

    def start_camera(self, source):
        try:
            self.camera = cv2.VideoCapture(source)
            if not self.camera.isOpened():
                messagebox.showerror("Error", "Gagal membuka kamera!")
                self.hist_window.destroy()
                return
            self.camera_running = True
            self.app.update_status("Kamera aktif - Mode Histogram")
            self.update_camera_loop()
        except Exception as e:
            messagebox.showerror("Error", f"Error: {str(e)}")
            self.hist_window.destroy()

    def update_camera_loop(self):
        if self.camera_running and self.camera:
            ret, frame = self.camera.read()
            if ret:
                self.current_frame = frame.copy()
                frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                img = Image.fromarray(cv2.resize(frame_rgb, (480, 360)))
                self.tk_img = ImageTk.PhotoImage(img)
                self.camera_canvas.delete("all")
                self.camera_canvas.create_image(240, 180, image=self.tk_img)
            self.hist_window.after(30, self.update_camera_loop)

    # ============== Core Functions ==============

    # ============== Core Functions ==============

    def capture_image(self):
        """Capture current frame or retake"""
        if not self.camera_running:
            # RETAKE MODE
            self.camera_running = True
            self.capture_btn.config(text="📷 Capture", bg="#2ecc71")
            self.histogram_btn.config(state=tk.DISABLED)
            self.save_btn.config(state=tk.DISABLED)
            self.captured_image = None
            self.update_camera_loop()
            self.app.update_status("Kamera live view")
        else:
            # CAPTURE MODE
            if self.camera and self.current_frame is not None:
                self.captured_image = self.current_frame.copy()
                self.camera_running = False  # Pause camera loop
                
                # Show captured image on canvas (freeze)
                frame_rgb = cv2.cvtColor(self.captured_image, cv2.COLOR_BGR2RGB)
                img = Image.fromarray(cv2.resize(frame_rgb, (480, 360)))
                self.tk_capture = ImageTk.PhotoImage(img)
                self.camera_canvas.delete("all")
                self.camera_canvas.create_image(240, 180, image=self.tk_capture)

                # Update buttons
                self.capture_btn.config(text="🔄 Ambil Ulang", bg="#f39c12")
                self.histogram_btn.config(state=tk.NORMAL)
                self.save_btn.config(state=tk.NORMAL)
                self.app.update_status("Gambar berhasil di-capture!")
                
                # Auto-generate histogram on capture? User request: "cameranya akan memunculkan hasil"
                # Let's wait for explicit Histogram click or just let it be.
            else:
                messagebox.showwarning("Peringatan", "Tidak ada frame kamera!")

    def show_histogram(self):
        """Calculate and display RGB histogram + statistics"""
        if self.captured_image is None:
            messagebox.showwarning("Peringatan", "Capture gambar terlebih dahulu!")
            return

        img_rgb = cv2.cvtColor(self.captured_image, cv2.COLOR_BGR2RGB)
        r_channel = img_rgb[:, :, 0].flatten()
        g_channel = img_rgb[:, :, 1].flatten()
        b_channel = img_rgb[:, :, 2].flatten()

        # Calculate component percentages
        sum_r = float(np.sum(r_channel))
        sum_g = float(np.sum(g_channel))
        sum_b = float(np.sum(b_channel))
        total_intensity = sum_r + sum_g + sum_b
        
        pct_r = (sum_r / total_intensity * 100) if total_intensity > 0 else 0
        pct_g = (sum_g / total_intensity * 100) if total_intensity > 0 else 0
        pct_b = (sum_b / total_intensity * 100) if total_intensity > 0 else 0
        
        # Update checkboxes text
        self.chk_r_btn.config(text=f"R ({pct_r:.1f}%)")
        self.chk_g_btn.config(text=f"G ({pct_g:.1f}%)")
        self.chk_b_btn.config(text=f"B ({pct_b:.1f}%)")

        # Store histogram data for refresh
        self._hist_data = {
            'r': r_channel, 'g': g_channel, 'b': b_channel
        }

        # Draw histogram
        self._draw_histogram()

        # Calculate statistics
        self.current_stats = {
            'skewness_r': float(stats.skew(r_channel)),
            'skewness_g': float(stats.skew(g_channel)),
            'skewness_b': float(stats.skew(b_channel)),
            'average_r': float(np.mean(r_channel)),
            'average_g': float(np.mean(g_channel)),
            'average_b': float(np.mean(b_channel)),
            'std_r': float(np.std(r_channel)),
            'std_g': float(np.std(g_channel)),
            'std_b': float(np.std(b_channel)),
            'kurtosis_r': float(stats.kurtosis(r_channel)),
            'kurtosis_g': float(stats.kurtosis(g_channel)),
            'kurtosis_b': float(stats.kurtosis(b_channel)),
        }

        # Update info panel
        h, w = self.captured_image.shape[:2]
        self.stat_labels['info'].config(text=f"{w} x {h} px | RGB")
        self.stat_labels['skewness'].config(
            text=f"R: {self.current_stats['skewness_r']:.4f}  "
                 f"G: {self.current_stats['skewness_g']:.4f}  "
                 f"B: {self.current_stats['skewness_b']:.4f}")
        self.stat_labels['average'].config(
            text=f"R: {self.current_stats['average_r']:.2f}  "
                 f"G: {self.current_stats['average_g']:.2f}  "
                 f"B: {self.current_stats['average_b']:.2f}")
        self.stat_labels['std'].config(
            text=f"R: {self.current_stats['std_r']:.4f}  "
                 f"G: {self.current_stats['std_g']:.4f}  "
                 f"B: {self.current_stats['std_b']:.4f}")
        self.stat_labels['kurtosis'].config(
            text=f"R: {self.current_stats['kurtosis_r']:.4f}  "
                 f"G: {self.current_stats['kurtosis_g']:.4f}  "
                 f"B: {self.current_stats['kurtosis_b']:.4f}")

        self.app.update_status("Histogram dan statistik berhasil dihitung!")

    def _draw_histogram(self):
        """Draw histogram on matplotlib canvas"""
        if not hasattr(self, '_hist_data'):
            return
        self.ax.clear()
        self.ax.set_facecolor('#1a252f')
        self.ax.tick_params(colors='white', labelsize=7)
        self.ax.spines['bottom'].set_color('white')
        self.ax.spines['left'].set_color('white')
        self.ax.spines['top'].set_visible(False)
        self.ax.spines['right'].set_visible(False)
        self.ax.set_xlabel('Intensitas', color='white', fontsize=8)
        self.ax.set_ylabel('Frekuensi', color='white', fontsize=8)

        if self.show_r.get():
            counts, bins, patches = self.ax.hist(self._hist_data['r'], bins=256, range=(0, 256),
                         color='red', alpha=0.4, label='R', histtype='stepfilled')
            # Annotate Peak
            max_val = counts.max()
            max_idx = np.argmax(counts)
            if max_val > 0:
                self.ax.annotate(f"{int(max_val)}", xy=(bins[max_idx], max_val), 
                                 xytext=(bins[max_idx], max_val + (max_val*0.05)),
                                 color='red', fontsize=7, ha='center',
                                 arrowprops=dict(arrowstyle="-", color='red', alpha=0.5))

        if self.show_g.get():
            counts, bins, patches = self.ax.hist(self._hist_data['g'], bins=256, range=(0, 256),
                         color='green', alpha=0.4, label='G', histtype='stepfilled')
            # Annotate Peak
            max_val = counts.max()
            max_idx = np.argmax(counts)
            if max_val > 0:
                self.ax.annotate(f"{int(max_val)}", xy=(bins[max_idx], max_val), 
                                 xytext=(bins[max_idx], max_val + (max_val*0.1)),
                                 color='green', fontsize=7, ha='center',
                                 arrowprops=dict(arrowstyle="-", color='green', alpha=0.5))

        if self.show_b.get():
            counts, bins, patches = self.ax.hist(self._hist_data['b'], bins=256, range=(0, 256),
                         color='blue', alpha=0.4, label='B', histtype='stepfilled')
            # Annotate Peak
            max_val = counts.max()
            max_idx = np.argmax(counts)
            if max_val > 0:
                self.ax.annotate(f"{int(max_val)}", xy=(bins[max_idx], max_val), 
                                 xytext=(bins[max_idx], max_val + (max_val*0.15)),
                                 color='#3498db', fontsize=7, ha='center',
                                 arrowprops=dict(arrowstyle="-", color='blue', alpha=0.5))

        if self.show_r.get() or self.show_g.get() or self.show_b.get():
            self.ax.legend(loc='upper right', fontsize=7, facecolor='#34495e',
                           edgecolor='white', labelcolor='white')
        
        self.fig.tight_layout()
        self.chart_canvas.draw()

    def _refresh_histogram(self):
        """Refresh histogram when checkboxes change"""
        if hasattr(self, '_hist_data'):
            self._draw_histogram()

    # ============== Database Methods ==============

    def save_to_database(self):
        """Save current statistics to database"""
        if not hasattr(self, 'current_stats'):
            messagebox.showwarning("Peringatan", "Tampilkan histogram terlebih dahulu!")
            return

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nama_citra = f"capture_{timestamp}"

        # Also save the image to gallery
        filepath = os.path.join(self.app.gallery_folder, f"{nama_citra}.png")
        if self.captured_image is not None:
            cv2.imwrite(filepath, self.captured_image)

        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        s = self.current_stats
        cursor.execute('''
            INSERT INTO histogram_data
            (nama_citra, skewness_r, skewness_g, skewness_b,
             average_r, average_g, average_b,
             std_r, std_g, std_b,
             kurtosis_r, kurtosis_g, kurtosis_b, timestamp)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (nama_citra,
              s['skewness_r'], s['skewness_g'], s['skewness_b'],
              s['average_r'], s['average_g'], s['average_b'],
              s['std_r'], s['std_g'], s['std_b'],
              s['kurtosis_r'], s['kurtosis_g'], s['kurtosis_b'],
              timestamp))
        conn.commit()
        conn.close()

        self._load_database()
        self.app.update_status(f"Data tersimpan: {nama_citra}")
        messagebox.showinfo("Sukses", f"Data histogram disimpan!\n{nama_citra}")

    def _load_database(self):
        """Load all data from database into treeview"""
        # Clear existing items
        for item in self.tree.get_children():
            self.tree.delete(item)

        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('SELECT id, nama_citra, skewness_r, skewness_g, skewness_b, '
                       'average_r, average_g, average_b, '
                       'std_r, std_g, std_b, '
                       'kurtosis_r, kurtosis_g, kurtosis_b FROM histogram_data ORDER BY id')
        rows = cursor.fetchall()
        conn.close()

        for idx, row in enumerate(rows, 1):
            db_id = row[0]
            nama = row[1]
            skew_text = f"R:{row[2]:.2f} G:{row[3]:.2f} B:{row[4]:.2f}"
            avg_text = f"R:{row[5]:.1f} G:{row[6]:.1f} B:{row[7]:.1f}"
            std_text = f"R:{row[8]:.2f} G:{row[9]:.2f} B:{row[10]:.2f}"
            kurt_text = f"R:{row[11]:.2f} G:{row[12]:.2f} B:{row[13]:.2f}"

            self.tree.insert("", tk.END, iid=str(db_id),
                             values=(idx, nama, skew_text, avg_text, std_text, kurt_text))

    def _on_tree_select(self, event):
        """Handle treeview row selection"""
        pass  # Could load selected image in future

    def delete_selected(self):
        """Delete selected row from database"""
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Peringatan", "Pilih data yang akan dihapus!")
            return

        if messagebox.askyesno("Konfirmasi", "Hapus data yang dipilih?"):
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            for item_id in selected:
                cursor.execute('DELETE FROM histogram_data WHERE id = ?', (int(item_id),))
            conn.commit()
            conn.close()
            self._load_database()
            self.app.update_status("Data berhasil dihapus")

    def export_to_excel(self):
        """Export database to Excel file"""
        try:
            import openpyxl
        except ImportError:
            messagebox.showerror("Error",
                                 "Module 'openpyxl' belum terinstall!\n"
                                 "Jalankan: pip install openpyxl")
            return

        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('SELECT id, nama_citra, skewness_r, skewness_g, skewness_b, '
                       'average_r, average_g, average_b, '
                       'std_r, std_g, std_b, '
                       'kurtosis_r, kurtosis_g, kurtosis_b, timestamp FROM histogram_data')
        rows = cursor.fetchall()
        conn.close()

        if not rows:
            messagebox.showwarning("Peringatan", "Tidak ada data untuk di-export!")
            return

        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            initialfile=f"histogram_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        if not filepath:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Histogram Data"

        # Headers
        headers = ["No", "Nama Citra",
                    "Skewness R", "Skewness G", "Skewness B",
                    "Average R", "Average G", "Average B",
                    "STD R", "STD G", "STD B",
                    "Kurtosis R", "Kurtosis G", "Kurtosis B",
                    "Timestamp"]
        ws.append(headers)

        # Style header
        from openpyxl.styles import Font, PatternFill, Alignment
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        for col_idx, cell in enumerate(ws[1], 1):
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for idx, row in enumerate(rows, 1):
            ws.append([idx] + list(row[1:]))

        # Auto-fit column width
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max(max_length + 2, 12)

        wb.save(filepath)
        self.app.update_status(f"Export berhasil: {os.path.basename(filepath)}")
        messagebox.showinfo("Sukses", f"Data berhasil di-export ke:\n{filepath}")

    # ============== Panel Control ==============

    def close_panel(self):
        """Close the histogram panel"""
        self.camera_running = False
        if self.camera:
            self.camera.release()
            self.camera = None
        self.hist_window.destroy()
        self.app.show_gallery_page()
